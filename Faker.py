from faker import Faker
from openpyxl import Workbook

faker = Faker() # Object creation
print(faker.name()) #trial with one name

# insert fake_test_data in workbook

wb = Workbook()
worksheet = wb.active

for i in range(1,10):
    worksheet.cell(row=i,column=1).value = faker.name()

# wb.save("Fake_test_data.xlsx")

new_sheet = wb.create_sheet('multiple_fakedata')

for rows in range(1,11): # for 10 rows
    for columns in range(1,4): # for 3 columns
        new_sheet.cell(row=rows, column=1).value = faker.name()
        new_sheet.cell(row=rows, column=2).value = faker.email()
        new_sheet.cell(row=rows, column=3).value = faker.password()

wb.save("Fake_test_data.xlsx")