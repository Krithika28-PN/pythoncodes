# first install openpyxl package --> pip install openpyxl
from openpyxl import Workbook


# wb = load_workbook('excelfile.xlsx') While opening an existing wb

# wb = Workbook()
# wb.save('excelfile.xlsx') While creating a newly wb

wb = Workbook() #create workbook object

worksheet = wb.active #active worksheet "Sheet_1"

testdata = [['Name','Surname'],['Rohit','Banik'],['Samarth','Kate'],['Sunita','Battul']]

for data in testdata:
    worksheet.append(data)
wb.save('excelfile.xlsx')

