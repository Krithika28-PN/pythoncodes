from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='excelfile.xlsx') #while opening a workbook use load_workbook
worksheet = wb.active

# print(worksheet['A2'].value)# to print specific value from sheet
# for printing whole sheet we need to use for loop for that we need row and col counts

row_count = worksheet.max_row
column_count = worksheet.max_column
print(row_count,column_count)

for rows in range(1,row_count+1):
    for columns in range(1,column_count+1):
        print(worksheet.cell(row=rows,column=columns).value, end=' ')
    print() # outside of the innerloop

